import csv
import io
from pathlib import Path

from openpyxl import load_workbook
from django import forms
from django.contrib import admin, messages
from django.contrib.auth.admin import UserAdmin as BaseUserAdmin
from django.contrib.auth.forms import UserChangeForm, UserCreationForm
from django.core.exceptions import ValidationError
from django.db import transaction
from django.http import HttpResponseRedirect
from django.shortcuts import redirect, render
from django.urls import path, reverse
from django.utils import timezone

from .forms import TicketGrantAdminForm
from .models import (
    CoachAvailability,
    CoachExpense,
    Court,
    FixedLesson,
    LessonWaitlist,
    LineAccountLink,
    Reservation,
    ScheduleSurveyResponse,
    ShopEstimateRequest,
    ShopProductMaster,
    STRINGING_COACH_NAMES,
    StringingOrder,
    TicketConsumption,
    TicketLedger,
    TicketPurchase,
    User,
    purchase_tickets,
)


admin.site.site_header = "Play Design Tennis 管理サイト"
admin.site.site_title = "Play Design Tennis 管理サイト"
admin.site.index_title = "管理メニュー"


_MODEL_VERBOSE_NAMES = {
    User: ("ユーザー", "ユーザー"),
    Court: ("コート", "コート"),
    CoachAvailability: ("コーチスケジュール", "コーチスケジュール"),
    FixedLesson: ("固定レッスン", "固定レッスン"),
    Reservation: ("予約", "予約"),
    TicketLedger: ("チケット履歴", "チケット履歴"),
    TicketPurchase: ("チケット購入", "チケット購入"),
    TicketConsumption: ("チケット消費", "チケット消費"),
    CoachExpense: ("経費", "経費"),
    StringingOrder: ("ガット張り依頼", "ガット張り依頼"),
    LineAccountLink: ("LINE連携", "LINE連携"),
    ScheduleSurveyResponse: ("時間帯アンケート回答", "時間帯アンケート回答"),
    ShopEstimateRequest: ("物販見積もり依頼", "物販見積もり依頼"),
    ShopProductMaster: ("Shop商品マスタ", "Shop商品マスタ"),
}


_FIELD_VERBOSE_NAMES = {
    User: {
        "username": "ログインID",
        "password": "パスワード",
        "full_name": "氏名",
        "email": "メールアドレス",
        "phone_number": "電話番号",
        "role": "権限種別",
        "contractor_hourly_wage": "業務委託コーチ時給",
        "member_level": "会員レベル",
        "ticket_balance": "チケット残数",
        "is_profile_completed": "会員情報入力済み",
        "is_staff": "管理画面利用可",
        "is_superuser": "管理者権限",
        "is_active": "有効",
        "last_login": "最終ログイン",
        "date_joined": "登録日時",
        "first_name": "名",
        "last_name": "姓",
        "groups": "グループ",
        "user_permissions": "ユーザー権限",
    },
    Court: {
        "name": "コート名",
        "is_active": "有効",
        "court_type": "コート種別",
    },
    CoachAvailability: {
        "coach": "主担当コーチ",
        "substitute_coach": "代行コーチ",
        "court": "コート",
        "lesson_type": "レッスン種別",
        "target_level": "対象レベル",
        "target_level_2": "対象レベル2",
        "start_at": "開始日時",
        "end_at": "終了日時",
        "capacity": "定員",
        "coach_count": "コーチ人数",
        "court_count": "必要コート数",
        "note": "メモ",
        "status": "公開状態",
        "custom_ticket_price": "イベント消費チケット",
        "custom_duration_hours": "イベント時間",
        "created_at": "作成日時",
    },
    FixedLesson: {
        "title": "レッスン名",
        "coach": "主担当コーチ",
        "coach_2": "追加コーチ1",
        "coach_3": "追加コーチ2",
        "court": "コート",
        "members": "固定参加メンバー",
        "lesson_type": "レッスン種別",
        "target_level": "対象レベル",
        "target_level_2": "対象レベル2",
        "start_date": "繰り返し開始日",
        "weekday": "曜日",
        "start_hour": "開始時刻",
        "capacity": "定員",
        "coach_count": "コーチ人数",
        "court_count": "必要コート数",
        "weeks_ahead": "作成する開催回数",
        "is_active": "有効",
        "note": "メモ",
        "created_at": "作成日時",
    },
    Reservation: {
        "user": "会員",
        "coach": "主担当コーチ",
        "substitute_coach": "代行コーチ",
        "court": "コート",
        "availability": "コーチスケジュール",
        "fixed_lesson": "固定レッスン",
        "is_fixed_entry": "固定参加",
        "lesson_type": "レッスン種別",
        "target_level": "対象レベル",
        "target_level_2": "対象レベル2",
        "requested_court_type": "希望コート種別",
        "requested_court_note": "希望コートメモ",
        "approved_court_note": "承認コートメモ",
        "start_at": "開始日時",
        "end_at": "終了日時",
        "tickets_used": "消費チケット",
        "ticket_consumed_at": "チケット消費日時",
        "ticket_refunded_at": "チケット返却日時",
        "status": "状態",
        "canceled_at": "キャンセル日時",
        "cancellation_reason": "キャンセル理由",
        "custom_ticket_price": "イベント消費チケット",
        "custom_duration_hours": "イベント時間",
        "created_at": "作成日時",
    },
    TicketLedger: {
        "user": "会員",
        "reservation": "予約",
        "fixed_lesson": "固定レッスン",
        "change_amount": "増減枚数",
        "balance_after": "処理後残数",
        "reason": "理由",
        "note": "メモ",
        "created_by": "処理者",
        "created_at": "作成日時",
    },
    TicketPurchase: {
        "user": "会員",
        "purchase_type": "購入種別",
        "total_tickets": "購入枚数",
        "remaining_tickets": "残枚数",
        "unit_price": "単価",
        "label": "表示名",
        "note": "メモ",
        "created_by": "処理者",
        "purchased_at": "購入日時",
        "created_at": "作成日時",
    },
    TicketConsumption: {
        "user": "会員",
        "purchase": "購入チケット",
        "reservation": "予約",
        "fixed_lesson": "固定レッスン",
        "tickets_used": "消費枚数",
        "unit_price_snapshot": "消費時単価",
        "refunded_at": "返却日時",
        "refund_note": "返却メモ",
        "created_at": "作成日時",
    },
    CoachExpense: {
        "expense_date": "経費日",
        "category": "カテゴリ",
        "amount": "金額",
        "note": "メモ",
        "created_by": "登録者",
        "created_at": "作成日時",
    },
    StringingOrder: {
        "user": "会員",
        "assigned_coach": "担当コーチ",
        "racket_name": "ラケット名",
        "string_name": "ガット名",
        "tension_lbs": "テンション",
        "delivery_requested": "デリバリー希望",
        "delivery_location": "デリバリー場所",
        "preferred_delivery_time": "希望日時",
        "base_price": "基本料金",
        "delivery_fee": "デリバリー料金",
        "status": "状態",
        "note": "メモ",
        "created_at": "作成日時",
        "updated_at": "更新日時",
    },
    LineAccountLink: {
        "user": "ユーザー",
        "line_user_id": "LINEユーザーID",
        "is_active": "有効",
        "linked_at": "連携日時",
        "last_event_at": "最終イベント日時",
    },
    ScheduleSurveyResponse: {
        "user": "会員",
        "selected_days": "参加しやすい曜日",
        "selected_weekday_time_slots": "平日の希望時間帯",
        "selected_weekend_time_slots": "土日の希望時間帯",
        "selected_lesson_types": "希望レッスン種別",
        "preferred_frequency": "希望頻度",
        "free_comment": "自由記入",
        "answered_at": "回答日時",
        "created_at": "作成日時",
        "updated_at": "更新日時",
    },
    ShopProductMaster: {
        "product_type": "商品種別",
        "category": "カテゴリ",
        "brand": "ブランド",
        "product_name": "商品名",
        "display_name": "表示名",
        "product_code": "品番",
        "official_price": "定価",
        "sale_price": "販売価格",
        "image_url": "画像URL",
        "product_url": "商品URL",
        "description": "説明",
        "is_active": "有効",
        "sort_order": "並び順",
        "created_at": "作成日時",
        "updated_at": "更新日時",
    },
    ShopEstimateRequest: {
        "user": "会員",
        "product_category": "商品カテゴリ",
        "brand": "ブランド",
        "main_keyword": "商品キーワード",
        "main_product_name": "商品名",
        "main_official_price": "商品定価",
        "grip_size": "グリップサイズ",
        "string_source": "ガット種別",
        "string_keyword": "ガットキーワード",
        "string_product_name": "ガット名",
        "string_official_price": "ガット定価",
        "request_stringing": "ガット張り希望",
        "request_delivery": "デリバリー希望",
        "tension_lbs": "テンション",
        "note": "メモ",
        "created_at": "作成日時",
        "updated_at": "更新日時",
    },
}


def _apply_japanese_admin_labels():
    for model, names in _MODEL_VERBOSE_NAMES.items():
        model._meta.verbose_name = names[0]
        model._meta.verbose_name_plural = names[1]

    for model, field_labels in _FIELD_VERBOSE_NAMES.items():
        for field_name, label in field_labels.items():
            try:
                model._meta.get_field(field_name).verbose_name = label
            except Exception:
                continue


_apply_japanese_admin_labels()


COACH_ROLE_VALUES = ("coach", "contractor_coach")


def coach_user_queryset():
    return User.objects.filter(role__in=COACH_ROLE_VALUES).order_by("full_name", "username", "id")


DATETIME_INPUT_FORMATS = [
    "%Y-%m-%dT%H:%M",
    "%Y-%m-%d %H:%M",
    "%Y/%m/%d %H:%M",
]


class AdminHourDateTimeInput(forms.DateTimeInput):
    input_type = "datetime-local"
    format = "%Y-%m-%dT%H:%M"


class ShopProductMasterImportForm(forms.Form):
    IMPORT_MODE_UPDATE = "update"
    IMPORT_MODE_REPLACE = "replace"

    IMPORT_MODE_CHOICES = (
        (IMPORT_MODE_UPDATE, "既存を残して更新 / 追加"),
        (IMPORT_MODE_REPLACE, "既存を全削除して入れ替え"),
    )

    upload_file = forms.FileField(label="取込ファイル")
    import_mode = forms.ChoiceField(
        label="取込方法",
        choices=IMPORT_MODE_CHOICES,
        initial=IMPORT_MODE_UPDATE,
    )
    default_is_active = forms.BooleanField(
        label="is_active が空欄の行は有効にする",
        required=False,
        initial=True,
    )


class CoachAvailabilityAdminForm(forms.ModelForm):
    start_at = forms.DateTimeField(
        label="開始日時",
        input_formats=DATETIME_INPUT_FORMATS,
        widget=AdminHourDateTimeInput(attrs={"step": 3600}),
    )
    end_at = forms.DateTimeField(
        label="終了日時",
        input_formats=DATETIME_INPUT_FORMATS,
        widget=AdminHourDateTimeInput(attrs={"step": 3600}),
    )

    class Meta:
        model = CoachAvailability
        fields = "__all__"

    def __init__(self, *args, **kwargs):
        super().__init__(*args, **kwargs)

        if self.instance and self.instance.pk:
            if getattr(self.instance, "start_at", None):
                self.initial["start_at"] = self.instance.start_at
            if getattr(self.instance, "end_at", None):
                self.initial["end_at"] = self.instance.end_at


class ReservationAdminForm(forms.ModelForm):
    start_at = forms.DateTimeField(
        label="開始日時",
        input_formats=DATETIME_INPUT_FORMATS,
        widget=AdminHourDateTimeInput(attrs={"step": 3600}),
    )
    end_at = forms.DateTimeField(
        label="終了日時",
        input_formats=DATETIME_INPUT_FORMATS,
        widget=AdminHourDateTimeInput(attrs={"step": 3600}),
    )

    class Meta:
        model = Reservation
        fields = "__all__"

    def __init__(self, *args, **kwargs):
        super().__init__(*args, **kwargs)

        coach_qs = coach_user_queryset()
        self.fields["coach"].queryset = coach_qs
        self.fields["substitute_coach"].queryset = coach_qs
        self.fields["substitute_coach"].required = False

        if self.instance and self.instance.pk:
            if getattr(self.instance, "start_at", None):
                self.initial["start_at"] = self.instance.start_at
            if getattr(self.instance, "end_at", None):
                self.initial["end_at"] = self.instance.end_at


class FixedLessonAdminForm(forms.ModelForm):
    class Meta:
        model = FixedLesson
        fields = "__all__"

    def __init__(self, *args, **kwargs):
        super().__init__(*args, **kwargs)

        coach_qs = coach_user_queryset()
        self.fields["coach"].queryset = coach_qs
        if "coach_2" in self.fields:
            self.fields["coach_2"].queryset = coach_qs
            self.fields["coach_2"].required = False
        if "coach_3" in self.fields:
            self.fields["coach_3"].queryset = coach_qs
            self.fields["coach_3"].required = False

        label_map = {
            "title": "レッスン名",
            "is_active": "有効",
            "lesson_type": "レッスン種別",
            "target_level": "対象レベル1",
            "target_level_2": "対象レベル2",
            "start_date": "繰り返し開始日",
            "weekday": "曜日",
            "start_hour": "開始時刻",
            "weeks_ahead": "作成する開催回数",
            "coach": "主担当コーチ",
            "coach_2": "追加コーチ1",
            "coach_3": "追加コーチ2",
            "court": "コート",
            "coach_count": "コーチ人数",
            "court_count": "必要コート数",
            "capacity": "定員",
            "members": "固定参加メンバー",
            "note": "メモ",
        }
        for field_name, label in label_map.items():
            if field_name in self.fields:
                self.fields[field_name].label = label

        help_text_map = {
            "start_date": "この日付以降の最初の該当曜日から、レッスンカレンダーに表示されます。",
            "weekday": "繰り返し開催する曜日を選択してください。",
            "weeks_ahead": "繰り返し開始日以降、何回分のレッスンをカレンダーに表示・予約作成するかを指定します。例：1 = 初回のみ、4 = 4回分。",
            "target_level_2": "2つ目の対象レベルがある場合のみ選択してください。例：対象レベル1=初級、対象レベル2=初中級。",
            "members": "ここに登録した会員は、今後の固定レッスン予約へ反映されます。外した会員の未来予約はキャンセル扱いになります。",
            "capacity": "一般レッスンはコーチ人数×6名で自動調整されます。",
            "coach_2": "複数コーチ開催時のみ選択してください。",
            "coach_3": "複数コーチ開催時のみ選択してください。",
        }
        for field_name, help_text in help_text_map.items():
            if field_name in self.fields:
                self.fields[field_name].help_text = help_text


class CoachExpenseAdminForm(forms.ModelForm):
    class Meta:
        model = CoachExpense
        fields = "__all__"

    def __init__(self, *args, **kwargs):
        super().__init__(*args, **kwargs)
        self.fields["created_by"].queryset = coach_user_queryset()
        self.fields["created_by"].required = False


class StringingOrderAdminForm(forms.ModelForm):
    class Meta:
        model = StringingOrder
        fields = "__all__"

    def __init__(self, *args, **kwargs):
        super().__init__(*args, **kwargs)
        if "assigned_coach" in self.fields:
            self.fields["assigned_coach"].queryset = User.objects.filter(
                role=User.ROLE_COACH,
                full_name__in=STRINGING_COACH_NAMES,
            ).order_by("full_name", "username", "id")
            self.fields["assigned_coach"].required = False
            self.fields["assigned_coach"].label = "担当コーチ"


class CustomUserCreationForm(UserCreationForm):
    class Meta(UserCreationForm.Meta):
        model = User
        fields = ("username", "full_name", "email", "phone_number", "role", "contractor_hourly_wage", "member_level")


class CustomUserChangeForm(UserChangeForm):
    class Meta(UserChangeForm.Meta):
        model = User
        fields = "__all__"


@admin.register(User)
class UserAdmin(BaseUserAdmin):
    form = CustomUserChangeForm
    add_form = CustomUserCreationForm

    list_display = (
        "id",
        "username",
        "full_name",
        "email",
        "phone_number",
        "role",
        "contractor_hourly_wage_display",
        "member_level",
        "ticket_balance",
        "is_profile_completed",
        "is_staff",
        "is_superuser",
    )
    list_filter = ("role", "member_level", "is_profile_completed", "is_staff", "is_superuser", "is_active")
    search_fields = ("username", "full_name", "email", "phone_number")
    ordering = ("id",)
    actions = ("grant_tickets_selected", "grant_single_ticket", "grant_set4_tickets")

    fieldsets = (
        (None, {"fields": ("username", "password")}),
        ("会員情報", {"fields": ("full_name", "phone_number", "email", "member_level", "ticket_balance", "is_profile_completed")}),
        ("業務委託コーチ設定", {"fields": ("contractor_hourly_wage",), "description": "権限種別が業務委託コーチの場合に、給与計算で使用する時給を円単位で入力します。通常コーチ・会員では0円のままで問題ありません。"}),
        ("個人情報", {"fields": ("first_name", "last_name")}),
        ("権限", {"fields": ("role", "is_active", "is_staff", "is_superuser", "groups", "user_permissions")}),
        ("重要な日付", {"fields": ("last_login", "date_joined")}),
    )

    add_fieldsets = (
        (
            None,
            {
                "classes": ("wide",),
                "fields": (
                    "username",
                    "full_name",
                    "email",
                    "phone_number",
                    "role",
                    "contractor_hourly_wage",
                    "member_level",
                    "password1",
                    "password2",
                    "is_staff",
                    "is_superuser",
                ),
            },
        ),
    )

    @admin.display(description="業務委託コーチ時給", ordering="contractor_hourly_wage")
    def contractor_hourly_wage_display(self, obj):
        try:
            return obj.contractor_hourly_wage_label()
        except Exception:
            value = int(getattr(obj, "contractor_hourly_wage", 0) or 0)
            return f"{value:,}円/時" if value > 0 else "-"

    def get_urls(self):
        urls = super().get_urls()
        custom_urls = [
            path(
                "grant-tickets/",
                self.admin_site.admin_view(self.grant_tickets_view),
                name="club_user_grant_tickets",
            ),
        ]
        return custom_urls + urls

    @admin.action(description="チケット付与（条件指定・複数人一括対応）")
    def grant_tickets_selected(self, request, queryset):
        selected_ids = list(queryset.values_list("pk", flat=True))
        if not selected_ids:
            self.message_user(request, "対象会員を選択してください。", level=messages.WARNING)
            return None

        ids_query = ",".join([str(pk) for pk in selected_ids])
        url = reverse("admin:club_user_grant_tickets")
        return HttpResponseRedirect(f"{url}?ids={ids_query}")

    def grant_tickets_view(self, request):
        raw_ids = (request.GET.get("ids") or request.POST.get("ids") or "").strip()
        selected_ids = [int(value) for value in raw_ids.split(",") if value.strip().isdigit()]
        queryset = User.objects.filter(pk__in=selected_ids).order_by("id")

        if not queryset.exists():
            self.message_user(request, "対象会員が見つかりません。", level=messages.WARNING)
            return redirect("admin:club_user_changelist")

        members = queryset.filter(role="member")
        skipped_users = list(queryset.exclude(role="member"))

        if request.method == "POST":
            form = TicketGrantAdminForm(request.POST)
            if form.is_valid():
                success_count = 0
                error_messages = []

                purchase_type = form.resolved_purchase_type()
                reason = form.resolved_reason()
                label = form.resolved_label()
                note = form.resolved_note()
                tickets = form.cleaned_data["tickets"]
                unit_price = form.cleaned_data["unit_price"]

                for user in members:
                    try:
                        purchase_tickets(
                            user=user,
                            tickets=tickets,
                            unit_price=unit_price,
                            purchase_type=purchase_type,
                            reason=reason,
                            note=note,
                            created_by=request.user,
                            label=label,
                        )
                        success_count += 1
                    except Exception as e:
                        error_messages.append(f"{user} の付与に失敗しました: {e}")

                if skipped_users:
                    skipped_names = "、".join([str(user) for user in skipped_users])
                    self.message_user(
                        request,
                        f"member 以外はスキップしました: {skipped_names}",
                        level=messages.WARNING,
                    )

                for message_text in error_messages:
                    self.message_user(request, message_text, level=messages.ERROR)

                if success_count:
                    self.message_user(
                        request,
                        f"{success_count}件の会員へチケットを付与しました。",
                        level=messages.SUCCESS,
                    )

                return redirect("admin:club_user_changelist")
        else:
            initial = {
                "tickets": 1,
                "unit_price": 4000,
                "label": "1枚券",
                "note": "管理画面から一括付与",
            }
            form = TicketGrantAdminForm(initial=initial)

        context = {
            **self.admin_site.each_context(request),
            "opts": self.model._meta,
            "title": "チケット付与（条件指定）",
            "form": form,
            "selected_users": queryset,
            "member_users": members,
            "skipped_users": skipped_users,
            "ids": raw_ids,
        }
        return render(request, "admin/club/user/grant_tickets.html", context)

    @admin.action(description="チケット1枚を付与する（1枚 4,000円）")
    def grant_single_ticket(self, request, queryset):
        count = 0
        for user in queryset:
            if user.role != "member":
                continue
            try:
                purchase_tickets(
                    user=user,
                    tickets=1,
                    unit_price=4000,
                    purchase_type=TicketPurchase.PURCHASE_TYPE_SINGLE,
                    reason=TicketLedger.REASON_PURCHASE_SINGLE,
                    note="管理画面から1枚付与",
                    created_by=request.user,
                    label="1枚券",
                )
                count += 1
            except Exception as e:
                self.message_user(request, f"{user} の付与に失敗しました: {e}", level=messages.ERROR)
        if count:
            self.message_user(request, f"{count}件の会員へチケット1枚を付与しました。", level=messages.SUCCESS)

    @admin.action(description="4枚セットを付与する（1枚あたり 3,500円）")
    def grant_set4_tickets(self, request, queryset):
        count = 0
        for user in queryset:
            if user.role != "member":
                continue
            try:
                purchase_tickets(
                    user=user,
                    tickets=4,
                    unit_price=3500,
                    purchase_type=TicketPurchase.PURCHASE_TYPE_SET4,
                    reason=TicketLedger.REASON_PURCHASE_SET4,
                    note="管理画面から4枚セット付与",
                    created_by=request.user,
                    label="4枚セット",
                )
                count += 1
            except Exception as e:
                self.message_user(request, f"{user} の付与に失敗しました: {e}", level=messages.ERROR)
        if count:
            self.message_user(request, f"{count}件の会員へ4枚セットを付与しました。", level=messages.SUCCESS)


@admin.register(Court)
class CourtAdmin(admin.ModelAdmin):
    list_display = ("id", "name", "court_type", "is_active")
    list_filter = ("court_type", "is_active")
    search_fields = ("name",)


@admin.register(CoachAvailability)
class CoachAvailabilityAdmin(admin.ModelAdmin):
    form = CoachAvailabilityAdminForm
    list_display = (
        "id",
        "coach",
        "court",
        "lesson_type",
        "target_level_admin",
        "coach_count",
        "court_count",
        "capacity",
        "start_at",
        "end_at",
    )
    list_filter = ("coach", "court", "lesson_type", "target_level", "target_level_2")
    search_fields = ("coach__username", "coach__full_name", "court__name")

    @admin.display(description="対象レベル", ordering="target_level")
    def target_level_admin(self, obj):
        if hasattr(obj, "target_level_display_label"):
            return obj.target_level_display_label()
        return obj.get_target_level_display()


@admin.register(FixedLesson)
class FixedLessonAdmin(admin.ModelAdmin):
    form = FixedLessonAdminForm
    list_display = (
        "id",
        "operation_status_admin",
        "weekday_display_admin",
        "start_hour_display_admin",
        "lesson_title_admin",
        "target_level_admin",
        "coach_names_admin",
        "court_display_admin",
        "member_count_admin",
        "capacity_status_admin",
        "start_date",
        "end_date_admin",
        "occurrence_count_admin",
        "future_reservation_count_admin",
        "future_waitlist_count_admin",
        "is_active",
    )
    list_display_links = ("id", "lesson_title_admin")
    list_editable = ("is_active",)
    list_filter = (
        "is_active",
        "weekday",
        "target_level",
        "target_level_2",
        "lesson_type",
        "coach",
        "coach_2",
        "coach_3",
        "court",
        "start_date",
    )
    search_fields = (
        "title",
        "coach__username",
        "coach__full_name",
        "coach_2__username",
        "coach_2__full_name",
        "coach_3__username",
        "coach_3__full_name",
        "court__name",
        "members__username",
        "members__full_name",
    )
    filter_horizontal = ("members",)
    actions = (
        "sync_selected_fixed_lessons",
        "activate_selected_fixed_lessons",
        "deactivate_selected_fixed_lessons",
    )
    list_per_page = 50
    save_on_top = True

    readonly_fields = (
        "operation_help_admin",
        "member_count_admin",
        "capacity_status_admin",
        "end_date_admin",
        "future_reservation_count_admin",
        "future_waitlist_count_admin",
    )

    fieldsets = (
        ("運用確認", {
            "fields": (
                "operation_help_admin",
                "member_count_admin",
                "capacity_status_admin",
                "end_date_admin",
                "future_reservation_count_admin",
                "future_waitlist_count_admin",
            ),
            "description": "保存後、固定参加メンバーの変更は今後の予約へ自動反映されます。",
        }),
        ("レッスン基本情報", {
            "fields": (
                "title",
                "is_active",
                "lesson_type",
                "target_level",
                "target_level_2",
            )
        }),
        ("開催曜日・時間", {
            "fields": (
                "start_date",
                "weekday",
                "start_hour",
                "weeks_ahead",
            ),
            "description": "繰り返し開始日以降、指定曜日に一致する日付から、指定した開催回数分だけレッスンカレンダーへ表示・予約生成します。例：作成する開催回数が1なら初回のみ、4なら4回分です。",
        }),
        ("担当コーチ・コート", {
            "fields": (
                "coach",
                "coach_2",
                "coach_3",
                "court",
            ),
            "description": "coach_2 / coach_3 に入っているコーチにも、コーチ用参加者一覧・スケジュールに表示されます。",
        }),
        ("定員・必要数", {
            "fields": (
                "coach_count",
                "court_count",
                "capacity",
            ),
            "description": "一般レッスンでは、選択したコーチ人数に応じて定員と必要コート数が自動調整されます。",
        }),
        ("固定参加メンバー", {
            "fields": (
                "members",
            ),
            "description": "固定参加メンバーを追加・削除すると、保存時に今後の固定レッスン予約へ反映されます。",
        }),
        ("メモ", {
            "fields": (
                "note",
            )
        }),
    )

    def get_queryset(self, request):
        return (
            super()
            .get_queryset(request)
            .select_related("coach", "coach_2", "coach_3", "court")
            .prefetch_related("members")
        )

    def _future_start(self):
        return timezone.now()

    def _end_date_for_obj(self, obj):
        try:
            repeat_start = obj.start_date or timezone.localdate()
            initial_offset = (int(obj.weekday) - repeat_start.weekday()) % 7
            if int(obj.weeks_ahead or 0) <= 0:
                return repeat_start + timezone.timedelta(days=initial_offset)
            return repeat_start + timezone.timedelta(days=initial_offset + (7 * (int(obj.weeks_ahead or 1) - 1)))
        except Exception:
            return None

    @admin.display(description="状態", ordering="is_active")
    def operation_status_admin(self, obj):
        if not obj.is_active:
            return "停止中"
        if not obj.court_id:
            return "コート未設定"
        if obj.member_count_for_admin() >= obj.effective_capacity():
            return "満員"
        return "運用中"

    @admin.display(description="曜日", ordering="weekday")
    def weekday_display_admin(self, obj):
        return obj.get_weekday_display()

    @admin.display(description="開始", ordering="start_hour")
    def start_hour_display_admin(self, obj):
        return f"{int(obj.start_hour or 0):02d}:00"

    @admin.display(description="レッスン名", ordering="title")
    def lesson_title_admin(self, obj):
        return obj.title or obj.get_lesson_type_display()

    @admin.display(description="対象レベル", ordering="target_level")
    def target_level_admin(self, obj):
        if hasattr(obj, "target_level_display_label"):
            return obj.target_level_display_label()
        return obj.get_target_level_display()

    @admin.display(description="担当コーチ")
    def coach_names_admin(self, obj):
        return obj.coach_display_names()

    @admin.display(description="コート")
    def court_display_admin(self, obj):
        return obj.court_display()

    @admin.display(description="固定メンバー")
    def member_count_admin(self, obj):
        return obj.member_count_for_admin()

    @admin.display(description="定員状況")
    def capacity_status_admin(self, obj):
        member_count = obj.member_count_for_admin()
        capacity = obj.effective_capacity()
        if member_count >= capacity:
            return f"{member_count}/{capacity}名（満員）"
        return f"{member_count}/{capacity}名"

    @admin.display(description="表示終了目安")
    def end_date_admin(self, obj):
        end_date = self._end_date_for_obj(obj)
        if not end_date:
            return "-"
        return end_date.strftime("%Y-%m-%d")

    @admin.display(description="作成する開催回数", ordering="weeks_ahead")
    def occurrence_count_admin(self, obj):
        try:
            count = max(int(obj.weeks_ahead or 1), 1)
        except Exception:
            count = 1
        return f"{count}回分"

    @admin.display(description="今後予約")
    def future_reservation_count_admin(self, obj):
        if not obj.pk:
            return 0
        return Reservation.objects.filter(
            fixed_lesson=obj,
            start_at__gte=self._future_start(),
            status=Reservation.STATUS_ACTIVE,
        ).count()

    @admin.display(description="今後キャンセル待ち")
    def future_waitlist_count_admin(self, obj):
        if not obj.pk:
            return 0
        return LessonWaitlist.objects.filter(
            fixed_lesson=obj,
            start_at__gte=self._future_start(),
            status=LessonWaitlist.STATUS_WAITING,
        ).count()

    @admin.display(description="運用メモ")
    def operation_help_admin(self, obj):
        if not obj or not obj.pk:
            return "保存後に、固定メンバー数・今後予約数・キャンセル待ち数が表示されます。"

        messages_for_admin = []
        if not obj.is_active:
            messages_for_admin.append("この固定レッスンは停止中です。レッスンカレンダーには表示されません。")
        if not obj.court_id:
            messages_for_admin.append("コートが未設定です。予約生成・カレンダー表示で問題になる可能性があります。")
        if obj.member_count_for_admin() > obj.effective_capacity():
            messages_for_admin.append("固定メンバー数が定員を超えています。定員またはメンバーを確認してください。")
        if not messages_for_admin:
            messages_for_admin.append("運用上の大きな警告はありません。")

        return " / ".join(messages_for_admin)

    def save_related(self, request, form, formsets, change):
        super().save_related(request, form, formsets, change)
        fixed_lesson = form.instance
        try:
            changed_count = fixed_lesson.sync_future_reservations(created_by=request.user)
            self.message_user(
                request,
                f"固定レッスンを保存しました。今後の予約への反映件数: {changed_count}件。",
                level=messages.SUCCESS,
            )
        except Exception as e:
            self.message_user(
                request,
                f"固定レッスンの今後予約への反映に失敗しました: {e}",
                level=messages.ERROR,
            )

    @admin.action(description="選択した固定レッスンの今後予約を生成・再同期する")
    def sync_selected_fixed_lessons(self, request, queryset):
        total = 0
        for fixed_lesson in queryset:
            try:
                total += fixed_lesson.sync_future_reservations(created_by=request.user)
            except Exception as e:
                self.message_user(request, f"{fixed_lesson} の同期に失敗しました: {e}", level=messages.ERROR)
        self.message_user(request, f"固定レッスン予約を {total} 件生成・更新しました。", level=messages.SUCCESS)

    @admin.action(description="選択した固定レッスンを有効にする")
    def activate_selected_fixed_lessons(self, request, queryset):
        updated = queryset.update(is_active=True)
        self.message_user(request, f"{updated}件の固定レッスンを有効にしました。", level=messages.SUCCESS)

    @admin.action(description="選択した固定レッスンを無効にする")
    def deactivate_selected_fixed_lessons(self, request, queryset):
        updated = queryset.update(is_active=False)
        self.message_user(request, f"{updated}件の固定レッスンを無効にしました。", level=messages.SUCCESS)


@admin.register(Reservation)
class ReservationAdmin(admin.ModelAdmin):
    form = ReservationAdminForm
    list_display = (
        "id",
        "user",
        "coach",
        "substitute_coach",
        "court",
        "lesson_type",
        "target_level_admin",
        "tickets_used",
        "start_at",
        "end_at",
        "status",
        "is_fixed_entry",
    )
    list_filter = ("status", "lesson_type", "target_level", "target_level_2", "coach", "substitute_coach", "court", "is_fixed_entry")
    @admin.display(description="対象レベル", ordering="target_level")
    def target_level_admin(self, obj):
        if hasattr(obj, "target_level_display_label"):
            return obj.target_level_display_label()
        return obj.get_target_level_display()

    search_fields = (
        "user__username",
        "user__full_name",
        "coach__username",
        "coach__full_name",
        "substitute_coach__username",
        "substitute_coach__full_name",
        "court__name",
        "cancellation_reason",
        "requested_court_note",
        "approved_court_note",
    )


@admin.register(TicketLedger)
class TicketLedgerAdmin(admin.ModelAdmin):
    list_display = ("id", "user", "change_amount", "balance_after", "reason", "created_by", "created_at")
    list_filter = ("reason", "created_at")
    search_fields = (
        "user__username",
        "user__full_name",
        "note",
        "reservation__court__name",
        "fixed_lesson__title",
    )
    autocomplete_fields = ("user", "reservation", "fixed_lesson", "created_by")


@admin.register(TicketPurchase)
class TicketPurchaseAdmin(admin.ModelAdmin):
    list_display = (
        "id",
        "user",
        "purchase_type",
        "unit_price",
        "total_tickets",
        "remaining_tickets",
        "label",
        "purchased_at",
    )
    list_filter = ("purchase_type", "unit_price", "purchased_at")
    search_fields = ("user__username", "user__full_name", "label", "note")
    autocomplete_fields = ("user", "created_by")


@admin.register(TicketConsumption)
class TicketConsumptionAdmin(admin.ModelAdmin):
    list_display = (
        "id",
        "user",
        "reservation",
        "purchase",
        "tickets_used",
        "unit_price_snapshot",
        "refunded_at",
        "created_at",
    )
    list_filter = ("unit_price_snapshot", "created_at", "refunded_at")
    search_fields = (
        "user__username",
        "user__full_name",
        "reservation__court__name",
        "purchase__label",
    )
    autocomplete_fields = ("user", "purchase", "reservation", "fixed_lesson")


@admin.register(CoachExpense)
class CoachExpenseAdmin(admin.ModelAdmin):
    form = CoachExpenseAdminForm
    list_display = ("id", "expense_date", "category", "amount", "note", "created_by", "created_at")
    list_filter = ("category", "expense_date")
    search_fields = ("note",)


@admin.register(LessonWaitlist)
class LessonWaitlistAdmin(admin.ModelAdmin):
    list_display = (
        "id",
        "user",
        "start_at",
        "end_at",
        "lesson_type",
        "target_level_admin",
        "coach",
        "substitute_coach",
        "court",
        "status",
        "created_at",
    )
    list_filter = ("status", "lesson_type", "target_level", "target_level_2", "coach", "court", "start_at")

    @admin.display(description="対象レベル", ordering="target_level")
    def target_level_admin(self, obj):
        if hasattr(obj, "target_level_display_label"):
            return obj.target_level_display_label()
        return obj.get_target_level_display()

    search_fields = (
        "user__username",
        "user__full_name",
        "coach__username",
        "coach__full_name",
        "substitute_coach__username",
        "substitute_coach__full_name",
        "court__name",
        "note",
    )
    autocomplete_fields = ("user", "coach", "substitute_coach", "court", "availability", "fixed_lesson")
    readonly_fields = ("created_at", "updated_at", "canceled_at", "converted_at")
    fieldsets = (
        ("キャンセル待ち情報", {
            "fields": (
                "user",
                "status",
                "lesson_type",
                "target_level",
                "target_level_2",
                "start_at",
                "end_at",
            )
        }),
        ("担当・場所", {
            "fields": (
                "coach",
                "substitute_coach",
                "court",
                "availability",
                "fixed_lesson",
            )
        }),
        ("メモ・日時", {
            "fields": (
                "note",
                "created_at",
                "updated_at",
                "canceled_at",
                "converted_at",
            )
        }),
    )


@admin.register(StringingOrder)
class StringingOrderAdmin(admin.ModelAdmin):
    form = StringingOrderAdminForm
    list_display = (
        "id",
        "user",
        "assigned_coach",
        "racket_name",
        "string_name",
        "tension_lbs",
        "delivery_requested",
        "base_price",
        "delivery_fee",
        "status",
        "created_at",
        "updated_at",
    )
    list_filter = ("status", "delivery_requested", "assigned_coach", "created_at")
    search_fields = (
        "user__username",
        "user__full_name",
        "assigned_coach__username",
        "assigned_coach__full_name",
        "racket_name",
        "string_name",
        "delivery_location",
        "preferred_delivery_time",
        "tension_lbs",
        "note",
    )
    autocomplete_fields = ("user",)
    readonly_fields = ("created_at", "updated_at")

    fieldsets = (
        ("依頼情報", {
            "fields": (
                "user",
                "assigned_coach",
                "status",
            )
        }),
        ("内容", {
            "fields": (
                "racket_name",
                "string_name",
                "delivery_requested",
                "tension_lbs",
                "delivery_location",
                "preferred_delivery_time",
                "note",
            )
        }),
        ("料金", {
            "fields": (
                "base_price",
                "delivery_fee",
            )
        }),
        ("日時", {
            "fields": (
                "created_at",
                "updated_at",
            )
        }),
    )


@admin.register(LineAccountLink)
class LineAccountLinkAdmin(admin.ModelAdmin):
    list_display = ("id", "user", "line_user_id", "is_active", "linked_at", "last_event_at")
    list_filter = ("is_active",)
    search_fields = ("user__username", "user__full_name", "line_user_id")


@admin.register(ScheduleSurveyResponse)
class ScheduleSurveyResponseAdmin(admin.ModelAdmin):
    list_display = (
        "id",
        "user",
        "preferred_frequency",
        "answered_at",
        "updated_at",
    )
    list_filter = ("preferred_frequency", "answered_at", "updated_at")
    search_fields = (
        "user__username",
        "user__full_name",
        "free_comment",
    )
    readonly_fields = ("answered_at", "created_at", "updated_at")


@admin.register(ShopProductMaster)
class ShopProductMasterAdmin(admin.ModelAdmin):
    change_list_template = "admin/club/shopproductmaster/change_list.html"

    list_display = (
        "id",
        "product_type",
        "category",
        "brand",
        "display_label_admin",
        "product_code",
        "official_price_display",
        "sale_price_display",
        "is_active",
        "sort_order",
        "updated_at",
    )
    list_filter = ("product_type", "category", "brand", "is_active")
    search_fields = (
        "product_name",
        "display_name",
        "product_code",
        "description",
        "spec_weight_unstrung",
        "spec_gauge",
    )
    readonly_fields = (
        "created_at",
        "updated_at",
        "sale_price_display",
        "racket_spec_summary",
        "string_spec_summary",
    )
    list_per_page = 100
    ordering = ("brand", "category", "product_type", "sort_order", "product_name", "id")

    fieldsets = (
        ("基本情報", {
            "fields": (
                "product_type",
                "category",
                "brand",
                "product_name",
                "display_name",
                "product_code",
                "official_price",
                "sale_price_display",
                "is_active",
                "sort_order",
            )
        }),
        ("リンク", {
            "fields": (
                "product_url",
                "image_url",
            )
        }),
        ("補足", {
            "fields": (
                "description",
            )
        }),
        ("ラケットスペック", {
            "fields": (
                "spec_weight_unstrung",
                "spec_string_pattern",
                "spec_head_size",
                "spec_balance",
                "spec_length",
                "spec_beam",
                "racket_spec_summary",
            )
        }),
        ("ガットスペック", {
            "fields": (
                "spec_gauge",
                "spec_set_length",
                "string_spec_summary",
            )
        }),
        ("日時", {
            "fields": (
                "created_at",
                "updated_at",
            )
        }),
    )

    @admin.display(description="表示名")
    def display_label_admin(self, obj):
        return obj.display_name or obj.product_name

    @admin.display(description="定価")
    def official_price_display(self, obj):
        return f"{int(obj.official_price or 0)}円"

    @admin.display(description="販売価格")
    def sale_price_display(self, obj):
        return f"{int(obj.sale_price or 0)}円"

    @admin.display(description="ラケットスペック")
    def racket_spec_summary(self, obj):
        return " / ".join(obj.racket_spec_lines()) or "-"

    @admin.display(description="ガットスペック")
    def string_spec_summary(self, obj):
        return " / ".join(obj.string_spec_lines()) or "-"

    @admin.display(description="業務委託コーチ時給", ordering="contractor_hourly_wage")
    def contractor_hourly_wage_display(self, obj):
        try:
            return obj.contractor_hourly_wage_label()
        except Exception:
            value = int(getattr(obj, "contractor_hourly_wage", 0) or 0)
            return f"{value:,}円/時" if value > 0 else "-"

    def get_urls(self):
        urls = super().get_urls()
        custom_urls = [
            path(
                "import-products/",
                self.admin_site.admin_view(self.import_products_view),
                name="club_shopproductmaster_import_products",
            ),
        ]
        return custom_urls + urls

    def import_products_view(self, request):
        if request.method == "POST":
            form = ShopProductMasterImportForm(request.POST, request.FILES)
            if form.is_valid():
                upload_file = form.cleaned_data["upload_file"]
                import_mode = form.cleaned_data["import_mode"]
                default_is_active = form.cleaned_data["default_is_active"]
                try:
                    result = self._import_uploaded_products(
                        upload_file=upload_file,
                        import_mode=import_mode,
                        default_is_active=default_is_active,
                    )
                    self.message_user(
                        request,
                        (
                            f"商品マスタ取込が完了しました。"
                            f" 追加: {result['created']}件 / 更新: {result['updated']}件 / "
                            f"スキップ: {result['skipped']}件"
                        ),
                        level=messages.SUCCESS,
                    )
                    if result["errors"]:
                        for error in result["errors"][:20]:
                            self.message_user(request, error, level=messages.WARNING)
                    return redirect("admin:club_shopproductmaster_changelist")
                except ValidationError as e:
                    for message_text in e.messages:
                        self.message_user(request, message_text, level=messages.ERROR)
                except Exception as e:
                    self.message_user(request, f"取込に失敗しました: {e}", level=messages.ERROR)
        else:
            form = ShopProductMasterImportForm()

        context = {
            **self.admin_site.each_context(request),
            "opts": self.model._meta,
            "title": "Shop商品マスタ一括取込",
            "form": form,
        }
        return render(request, "admin/club/shopproductmaster/import_products.html", context)

    def _import_uploaded_products(self, *, upload_file, import_mode, default_is_active):
        rows = self._read_uploaded_rows(upload_file)
        normalized_rows = self._normalize_import_rows(rows, default_is_active=default_is_active)

        created_count = 0
        updated_count = 0
        skipped_count = 0
        errors = []

        with transaction.atomic():
            if import_mode == ShopProductMasterImportForm.IMPORT_MODE_REPLACE:
                ShopProductMaster.objects.all().delete()

            for index, row in enumerate(normalized_rows, start=2):
                try:
                    instance = self._find_existing_product(row)
                    if instance is None:
                        instance = ShopProductMaster()

                    instance.product_type = row["product_type"]
                    instance.category = row["category"]
                    instance.brand = row["brand"]
                    instance.product_name = row["product_name"]
                    instance.display_name = row["display_name"]
                    instance.product_code = row["product_code"]
                    instance.official_price = row["official_price"]
                    instance.image_url = row["image_url"]
                    instance.product_url = row["product_url"]
                    instance.description = row["description"]
                    instance.spec_weight_unstrung = row["spec_weight_unstrung"]
                    instance.spec_string_pattern = row["spec_string_pattern"]
                    instance.spec_head_size = row["spec_head_size"]
                    instance.spec_balance = row["spec_balance"]
                    instance.spec_length = row["spec_length"]
                    instance.spec_beam = row["spec_beam"]
                    instance.spec_gauge = row["spec_gauge"]
                    instance.spec_set_length = row["spec_set_length"]
                    instance.sort_order = row["sort_order"]
                    instance.is_active = row["is_active"]
                    instance.full_clean()
                    is_update = bool(instance.pk)
                    instance.save()

                    if is_update:
                        updated_count += 1
                    else:
                        created_count += 1
                except ValidationError as e:
                    skipped_count += 1
                    joined = " / ".join(e.messages)
                    errors.append(f"{index}行目をスキップしました: {joined}")
                except Exception as e:
                    skipped_count += 1
                    errors.append(f"{index}行目をスキップしました: {e}")

        return {
            "created": created_count,
            "updated": updated_count,
            "skipped": skipped_count,
            "errors": errors,
        }

    def _read_uploaded_rows(self, upload_file):
        suffix = Path(upload_file.name).suffix.lower()

        if suffix == ".csv":
            decoded = upload_file.read().decode("utf-8-sig")
            reader = csv.DictReader(io.StringIO(decoded))
            return list(reader)

        if suffix == ".xlsx":
            workbook = load_workbook(upload_file, data_only=True)
            sheet = self._pick_product_master_sheet(workbook)
            values = list(sheet.values)
            if not values:
                return []
            headers = [str(value).strip() if value is not None else "" for value in values[0]]
            rows = []
            for row_values in values[1:]:
                row_dict = {}
                for idx, header in enumerate(headers):
                    if not header:
                        continue
                    row_dict[header] = row_values[idx] if idx < len(row_values) else None
                rows.append(row_dict)
            return rows

        raise ValidationError("取込できるのは .csv または .xlsx ファイルのみです。")

    def _pick_product_master_sheet(self, workbook):
        preferred_names = [
            "商品マスタ_全件",
            "商品マスタ",
            "products",
            "product_master",
        ]
        expected_headers = {
            "product_type",
            "category",
            "brand",
            "product_name",
            "display_name",
            "product_code",
            "official_price",
        }

        for sheet_name in preferred_names:
            if sheet_name in workbook.sheetnames:
                sheet = workbook[sheet_name]
                headers = self._sheet_header_set(sheet)
                if expected_headers.intersection(headers):
                    return sheet

        best_sheet = None
        best_score = -1
        for sheet in workbook.worksheets:
            headers = self._sheet_header_set(sheet)
            score = len(expected_headers.intersection(headers))
            if score > best_score:
                best_score = score
                best_sheet = sheet

        if best_sheet is None or best_score <= 0:
            raise ValidationError(
                "Excel内に商品マスタの明細シートが見つかりません。"
                " 『商品マスタ_全件』シート、または product_name / brand / category などの列を含むシートを使ってください。"
            )

        return best_sheet

    def _sheet_header_set(self, sheet):
        header_values = []
        for cell_value in next(sheet.iter_rows(min_row=1, max_row=1, values_only=True), []):
            if cell_value is None:
                continue
            header_values.append(str(cell_value).strip())
        return set(header_values)

    def _normalize_import_rows(self, rows, *, default_is_active):
        brand_map = {
            "yonex": ShopProductMaster.BRAND_YONEX,
            "YONEX": ShopProductMaster.BRAND_YONEX,
            "wilson": ShopProductMaster.BRAND_WILSON,
            "Wilson": ShopProductMaster.BRAND_WILSON,
            "babolat": ShopProductMaster.BRAND_BABOLAT,
            "Babolat": ShopProductMaster.BRAND_BABOLAT,
            "head": ShopProductMaster.BRAND_HEAD,
            "HEAD": ShopProductMaster.BRAND_HEAD,
            "prince": ShopProductMaster.BRAND_PRINCE,
            "Prince": ShopProductMaster.BRAND_PRINCE,
            "dunlop": ShopProductMaster.BRAND_DUNLOP,
            "DUNLOP": ShopProductMaster.BRAND_DUNLOP,
            "tecnifibre": ShopProductMaster.BRAND_TECHNIFIBRE,
            "Tecnifibre": ShopProductMaster.BRAND_TECHNIFIBRE,
            "other": ShopProductMaster.BRAND_OTHER,
            "その他": ShopProductMaster.BRAND_OTHER,
        }
        category_map = {
            "racket": ShopProductMaster.CATEGORY_RACKET,
            "ラケット": ShopProductMaster.CATEGORY_RACKET,
            "string": ShopProductMaster.CATEGORY_STRING,
            "ガット": ShopProductMaster.CATEGORY_STRING,
            "アクセサリ": ShopProductMaster.CATEGORY_ACCESSORY,
            "accessory": ShopProductMaster.CATEGORY_ACCESSORY,
        }
        product_type_map = {
            "main": ShopProductMaster.PRODUCT_TYPE_MAIN,
            "メイン商品": ShopProductMaster.PRODUCT_TYPE_MAIN,
            "string": ShopProductMaster.PRODUCT_TYPE_STRING,
            "ガット": ShopProductMaster.PRODUCT_TYPE_STRING,
        }

        def pick(row, *keys):
            for key in keys:
                if key in row and row[key] not in (None, ""):
                    return row[key]
            return ""

        def clean_text(value):
            return str(value or "").strip()

        def clean_int(value, default=0):
            if value in (None, ""):
                return default
            text = str(value).replace(",", "").strip()
            return int(text or default)

        def clean_bool(value, default=True):
            if value in (None, ""):
                return default
            text = str(value).strip().lower()
            if text in ("1", "true", "yes", "y", "on", "有効"):
                return True
            if text in ("0", "false", "no", "n", "off", "無効"):
                return False
            return default

        normalized = []
        for row in rows:
            product_name = clean_text(
                pick(row, "product_name", "商品名", "name", "品名", "モデル名")
            )
            if not product_name:
                continue

            brand_raw = clean_text(pick(row, "brand", "ブランド"))
            category_raw = clean_text(pick(row, "category", "カテゴリ", "category_name"))
            product_type_raw = clean_text(pick(row, "product_type", "商品種別", "type"))

            product_type = product_type_map.get(product_type_raw, ShopProductMaster.PRODUCT_TYPE_MAIN)
            category = category_map.get(category_raw, ShopProductMaster.CATEGORY_RACKET)
            brand = brand_map.get(brand_raw, ShopProductMaster.BRAND_OTHER)

            normalized.append(
                {
                    "product_type": product_type,
                    "category": ShopProductMaster.CATEGORY_STRING if product_type == ShopProductMaster.PRODUCT_TYPE_STRING else category,
                    "brand": brand,
                    "product_name": product_name,
                    "display_name": clean_text(pick(row, "display_name", "表示名")),
                    "product_code": clean_text(pick(row, "product_code", "品番", "商品コード", "code")),
                    "official_price": clean_int(pick(row, "official_price", "定価", "list_price", "price"), 0),
                    "image_url": clean_text(pick(row, "image_url", "画像URL")),
                    "product_url": clean_text(pick(row, "product_url", "商品URL", "url")),
                    "description": clean_text(pick(row, "description", "説明", "備考")),
                    "spec_weight_unstrung": clean_text(pick(row, "spec_weight_unstrung", "重量（ガット無し）", "重量(ガット無し)", "重量")),
                    "spec_string_pattern": clean_text(pick(row, "spec_string_pattern", "ストリングパターン")),
                    "spec_head_size": clean_text(pick(row, "spec_head_size", "ヘッドサイズ")),
                    "spec_balance": clean_text(pick(row, "spec_balance", "バランス")),
                    "spec_length": clean_text(pick(row, "spec_length", "長さ")),
                    "spec_beam": clean_text(pick(row, "spec_beam", "ビーム")),
                    "spec_gauge": clean_text(pick(row, "spec_gauge", "ゲージ")),
                    "spec_set_length": clean_text(pick(row, "spec_set_length", "セット長", "長さ(ガット)", "ガット長さ")),
                    "sort_order": clean_int(pick(row, "sort_order", "並び順"), 0),
                    "is_active": clean_bool(pick(row, "is_active", "公開", "active"), default_is_active),
                }
            )
        return normalized

    def _find_existing_product(self, row):
        product_code = (row.get("product_code") or "").strip()
        if product_code:
            return ShopProductMaster.objects.filter(product_code=product_code).first()

        return ShopProductMaster.objects.filter(
            brand=row["brand"],
            category=row["category"],
            product_type=row["product_type"],
            product_name=row["product_name"],
        ).first()


@admin.register(ShopEstimateRequest)
class ShopEstimateRequestAdmin(admin.ModelAdmin):
    list_display = (
        "id",
        "created_at",
        "user_display",
        "handling_status",
        "product_category",
        "brand",
        "main_product_display",
        "main_sale_price_display",
        "string_source",
        "string_sale_price_display",
        "stringing_fee_display",
        "estimated_total_display",
    )
    list_filter = (
        "handling_status",
        "product_category",
        "brand",
        "string_source",
        "request_stringing",
        "created_at",
    )
    search_fields = (
        "user__username",
        "user__full_name",
        "main_keyword",
        "main_product_name",
        "string_keyword",
        "string_product_name",
        "note",
        "admin_note",
    )
    readonly_fields = (
        "created_at",
        "updated_at",
        "main_sale_price_display",
        "string_sale_price_display",
        "stringing_fee_display",
        "estimated_total_display",
    )
    list_select_related = ("user",)
    list_per_page = 50
    date_hierarchy = "created_at"
    actions = (
        "mark_as_checked",
        "mark_as_ordered",
        "mark_as_completed",
        "mark_as_canceled",
    )

    fieldsets = (
        ("基本情報", {
            "fields": (
                "user",
                "handling_status",
                "product_category",
                "brand",
                "note",
                "admin_note",
            )
        }),
        ("メイン商品", {
            "fields": (
                "main_keyword",
                "main_product_name",
                "main_official_price",
                "main_sale_price_display",
            )
        }),
        ("ガット", {
            "fields": (
                "string_source",
                "string_keyword",
                "string_product_name",
                "string_official_price",
                "string_sale_price_display",
            )
        }),
        ("ガット張り", {
            "fields": (
                "request_stringing",
                "tension_lbs",
                "stringing_fee_display",
            )
        }),
        ("見積もり", {
            "fields": (
                "estimated_total_display",
            )
        }),
        ("日時", {
            "fields": (
                "created_at",
                "updated_at",
            )
        }),
    )

    @admin.display(description="会員")
    def user_display(self, obj):
        try:
            return obj.user.display_name()
        except Exception:
            return str(obj.user)

    @admin.display(description="商品")
    def main_product_display(self, obj):
        return obj.main_product_name or obj.main_keyword or "-"

    @admin.display(description="メイン販売価格")
    def main_sale_price_display(self, obj):
        return f"{obj.main_sale_price}円"

    @admin.display(description="ガット販売価格")
    def string_sale_price_display(self, obj):
        return f"{obj.string_sale_price}円"

    @admin.display(description="張り料金")
    def stringing_fee_display(self, obj):
        return f"{obj.stringing_fee}円"

    @admin.display(description="見積合計")
    def estimated_total_display(self, obj):
        return f"{obj.estimated_total}円"

    @admin.action(description="選択した物販申込を『確認済み』にする")
    def mark_as_checked(self, request, queryset):
        updated = queryset.exclude(
            handling_status=ShopEstimateRequest.HANDLING_STATUS_CHECKED
        ).update(handling_status=ShopEstimateRequest.HANDLING_STATUS_CHECKED)
        self.message_user(request, f"{updated}件を確認済みに更新しました。", level=messages.SUCCESS)

    @admin.action(description="選択した物販申込を『発注済み』にする")
    def mark_as_ordered(self, request, queryset):
        updated = queryset.exclude(
            handling_status=ShopEstimateRequest.HANDLING_STATUS_ORDERED
        ).update(handling_status=ShopEstimateRequest.HANDLING_STATUS_ORDERED)
        self.message_user(request, f"{updated}件を発注済みに更新しました。", level=messages.SUCCESS)

    @admin.action(description="選択した物販申込を『対応完了』にする")
    def mark_as_completed(self, request, queryset):
        updated = queryset.exclude(
            handling_status=ShopEstimateRequest.HANDLING_STATUS_COMPLETED
        ).update(handling_status=ShopEstimateRequest.HANDLING_STATUS_COMPLETED)
        self.message_user(request, f"{updated}件を対応完了に更新しました。", level=messages.SUCCESS)

    @admin.action(description="選択した物販申込を『キャンセル』にする")
    def mark_as_canceled(self, request, queryset):
        updated = queryset.exclude(
            handling_status=ShopEstimateRequest.HANDLING_STATUS_CANCELED
        ).update(handling_status=ShopEstimateRequest.HANDLING_STATUS_CANCELED)
        self.message_user(request, f"{updated}件をキャンセルに更新しました。", level=messages.SUCCESS)
